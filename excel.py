# -*- coding: utf-8 -*-
"""
Created on Mon Dec 26 11:38:06 2016

@author: Link
"""
from itertools import islice
from openpyxl.styles import Border, Side, Font, Alignment
from pandas import DataFrame
from openpyxl import load_workbook
from openpyxl import Workbook
#from openpyxl.utils.dataframe import dataframe_to_rows
from fontAndAlign import set_align
from drawHead import drawHead
from drawTail import drawTail
from appendRowsViaDataFrame import appendRowsViaDataFrame
#import numpy as np
#if len(sys.argv)>1:
#    fileName = sys.argv[1]
#    print sys.argv[1]
#else:
#    fileName = '20161205.xlsx'
#print u'请输入文件名或是将文件拖入窗口，按回车执行，会生成一个文件名+New.xlsx的文件'
#fileName = raw_input("") 
#wb2 = load_workbook(fileName)
fileName = '20170102.xlsx'
wb2 = load_workbook(fileName)


print wb2.get_sheet_names()
ws1 = wb2.active
data = ws1.values
cols = next(data)[1:]
data = list(data)
idx = [r[0] for r in data]
data = (islice(r, 1, None) for r in data)
df = DataFrame(data, index=idx, columns=cols)
#df = DataFrame(data)
#格式
borderNone = Border()
border = Border(left=Side(style='medium',color='FF000000'),right=Side(style='medium',color='FF000000'),top=Side(style='medium',color='FF000000'),bottom=Side(style='medium',color='FF000000'),diagonal=Side(style='medium',color='FF000000'),diagonal_direction=0,outline=Side(style='medium',color='FF000000'),vertical=Side(style='medium',color='FF000000'),horizontal=Side(style='medium',color='FF000000'))
fontObj2 = Font(size=9, italic=False)
fontObj3 = Font(size=14, italic=False, bold=True)
align = Alignment(horizontal='center', vertical='center', wrapText = True) 

# new workBook
wb = Workbook()
ws = []

group=[]
j = 0
init = 1
pastLayer = 0
for i in df.iloc[:,2]:
    if i < pastLayer  or init == 1: #本次层数小于上一次层数， 则标记为新一层的开始  或是初始状态
        group.append(j)
        init = 0
#    elif i != 1:
#        newGruopflag = 0
    pastLayer = i                         #更新上次层index                
    j = j+1                
groupNum = len(group)    

#split the dataframe
groupInList = []

if groupNum != 1: # Need to be split
    for i in range(0,groupNum):
        if i != groupNum-1:
            groupInList.append(df.iloc[group[i]:group[i+1]])
        else:
            groupInList.append(df.iloc[group[i]:len(df)])
else:
    groupInList.append(df)
#data process in each group
j = 0
for groupDf in groupInList:
    if j ==0:
        ws.append(wb.active)
    else:
        ws.append(wb.create_sheet("Sheet"))
    drawHead(ws[j])
    
    
    appendRowsViaDataFrame(ws[j], groupDf)
              
    drawTail(ws[j])  
    ws[j].row_dimensions[ws[j].max_row-1].height = 25
    ws[j].row_dimensions[1].height = 18
    set_align(ws[j],'A1:'+'G1',align,fontObj3,borderNone)
    set_align(ws[j],'A2:'+'G2',align,fontObj2,borderNone)
    set_align(ws[j],'A3:'+'G'+str(ws[j].max_row),align,fontObj2,border)
    j=j+1 

#save file    
fileNameSave=fileName[:-5]
wb.save(fileNameSave+'New'+'.xlsx')
print 'file has been created'